import streamlit as st
import pandas as pd
import easyocr
import re
from PIL import Image
import io
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImg

# 页面配置：开启局域网访问，手机同WiFi可打开
st.set_page_config(
    page_title="订单截图自动填表+嵌入原图工具",
    layout="wide",
    server_address="0.0.0.0",
    server_port=8501
)
st.title("🧾 订单截图自动填表（图片嵌入单元格）")

# 初始化OCR（缓存仅加载一次）
@st.cache_resource
def load_ocr():
    return easyocr.Reader(['ch_sim', 'en'], gpu=False, model_storage_directory="./.ocr_models")

reader = load_ocr()

# 侧边栏上传文件
st.sidebar.header("1. 上传门店总表(.xlsx)")
main_file = st.sidebar.file_uploader("上传你的湘大厨单据总表", type=["xlsx"])

st.sidebar.header("2. 批量上传送货单截图")
img_files = st.sidebar.file_uploader("多选多张图片", type=["png", "jpg", "jpeg"], accept_multiple_files=True)

# 门店纠错映射（修正OCR识别错字）
STORE_MAP = {
    "万象天地": "万象天地",
    "龙岗COCOPARK": "龙岗COCOPARK",
    "深业上城": "深业上城",
    "中航君尚": "中航君尚店",
    "欢乐颂": "欢乐颂",
    "万象食家": "万象食家",
    "怀德万象汇": "怀德万象汇",
    "星河雅宝": "星河雅宝",
    "中洲湾": "中洲湾",
    "布吉万象汇": "布吉万象汇",
    "深业东岭": "深业东岭",
    "卓悦中心": "卓悦中心",
    "中心书城": "中心书城",
    "红山": "红山",
    "天安云谷": "天安云谷",
    "东莞松山湖": "东莞松山湖",
    "方大城": "方大城",
    "欢乐港湾": "欢乐港湾",
    "前海科兴": "前海科兴",
    "壹方天地": "壹方天地",
    "大运天地": "大运天地",
    "雪花店": "雪花店",
    "后海汇": "后海汇",
    "绿景虹湾": "绿景虹湾",
    "天河城": "天河城",
    "万科": "万科"
}

def extract_info(text_list):
    """提取门店、日期、白玉丝瓜数量，增强容错"""
    full_text = " ".join(text_list)
    date_str = None
    store_name = None
    quantity = None

    # 兼容两种日期格式：2026-06-18 / 2026/06/18 / 6月18日
    date_match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2})', full_text)
    date_match2 = re.search(r'(\d{1,2})月(\d{1,2})日', full_text)
    if date_match:
        date_str = date_match.group(1).replace("/", "-")
    elif date_match2:
        m = int(date_match2.group(1))
        d = int(date_match2.group(2))
        date_str = f"2026-{m:02d}-{d:02d}"

    # 门店精准匹配
    for key in STORE_MAP.keys():
        if key in full_text:
            store_name = STORE_MAP[key]
            break
    if not store_name:
        parts = full_text.split("-")
        if len(parts) > 1:
            candidate = parts[-1].strip().replace("店", "")
            for key in STORE_MAP.keys():
                if key in candidate:
                    store_name = STORE_MAP[key]
                    break
            if not store_name:
                store_name = candidate

    # 提取白玉丝瓜数量
    for i, line in enumerate(text_list):
        if "白玉丝瓜" in line:
            for j in range(i, min(i+5, len(text_list))):
                nums = re.findall(r'\b(\d+)\b', text_list[j])
                for n in nums:
                    num = int(n)
                    if 0 < num < 1000:
                        quantity = num
                        break
                if quantity:
                    break
    if not quantity:
        match = re.search(r'白玉丝瓜.*?(\d+)\s*斤', full_text)
        if match:
            quantity = int(match.group(1))
    return store_name, date_str, quantity

# 核心函数：写入数字 + 把原图嵌入对应单元格
def fill_excel_with_image(excel_bytes, img_bytes, store, date, qty):
    # 读取原始表格
    wb = load_workbook(io.BytesIO(excel_bytes))
    sheet = wb["6月"]
    df = pd.read_excel(io.BytesIO(excel_bytes), header=1)
    df.columns = df.columns.str.strip()
    df['门店'] = df['门店'].fillna(method='ffill')

    # 解析日期列名
    dt = pd.to_datetime(date)
    col_name = f"{dt.month}月{dt.day}日"
    mask = df['门店'].str.contains(store, na=False)
    if not mask.any():
        return None, f"❌ 表格无门店：{store}"
    row_idx_df = df[mask].index[0]
    excel_row = row_idx_df + 3  # 适配表格表头偏移，根据你的表格微调

    # 定位日期列
    if col_name not in df.columns:
        return None, f"⚠️ 表格无日期列：{col_name}"
    excel_col = list(df.columns).index(col_name) + 1

    # 1. 写入数量数字
    sheet.cell(row=excel_row, column=excel_col, value=qty)
    # 2. 读取图片，嵌入当前日期单元格
    temp_img = io.BytesIO(img_bytes)
    pic = ExcelImg(temp_img)
    pic.width = 200
    pic.height = 130
    sheet.add_image(pic, sheet.cell(excel_row, excel_col).coordinate)

    # 导出新表格字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, f"✅ {store} | {col_name} | 数量{qty}，单据图片已嵌入单元格"

# 主逻辑
if main_file and img_files:
    raw_excel = main_file.getvalue()
    df_preview = pd.read_excel(io.BytesIO(raw_excel), header=1)
    df_preview.columns = df_preview.columns.str.strip()
    df_preview['门店'] = df_preview['门店'].fillna(method='ffill')

    st.subheader("📋 原始总表预览")
    st.dataframe(df_preview, height=300)

    if st.button("🚀 批量识别、填数字+嵌入原图"):
        progress = st.progress(0)
        log_list = []
        final_excel_data = raw_excel

        for idx, img_file in enumerate(img_files):
            # 读取图片二进制
            img_raw = img_file.getvalue()
            img_pil = Image.open(io.BytesIO(img_raw))
            img_np = np.array(img_pil)
            # OCR识别文字
            text_lines = reader.readtext(img_np, detail=0)
            shop, day_date, num = extract_info(text_lines)

            if shop and day_date and num:
                final_excel_data, msg = fill_excel_with_image(final_excel_data, img_raw, shop, day_date, num)
                if not final_excel_data:
                    log_list.append(msg)
                else:
                    log_list.append(msg)
            else:
                log_list.append(f"⚠️ 识别信息不全：门店={shop} 日期={day_date} 数量={num}")
            progress.progress((idx+1)/len(img_files))
        progress.empty()

        # 展示日志
        st.subheader("📝 批量处理日志")
        for line in log_list:
            st.write(line)

        # 下载成品表格
        st.download_button(
            label="📥 下载已填好+嵌入单据图的Excel",
            data=final_excel_data,
            file_name="湘大厨_已填单据.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # 预览填充后表格
        filled_df = pd.read_excel(io.BytesIO(final_excel_data), header=1)
        filled_df.columns = filled_df.columns.str.strip()
        st.subheader("📊 填充完成表格预览")
        st.dataframe(filled_df, height=300)

else:
    st.info("👈 左侧上传【门店总表Excel】+【送货单截图】才能开始识别")
    st.markdown("### 📱 手机使用方法（同WiFi）")
    st.code("""
1. 电脑运行：streamlit run ocr_fill_table.py
2. cmd输入ipconfig 查看电脑局域网IP（如192.168.1.50）
3. 手机浏览器打开：http://192.168.1.50:8501
4. 直接从手机相册上传送货单图片批量处理
""")
